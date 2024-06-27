from .worker_service_comparison import WorkerServiceComparison
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from io import BytesIO
from django.http import HttpResponse
from datetime import datetime


class WorkerServiceComparisonExcel:
    def __init__(self):
        self.worker_comparison = WorkerServiceComparison()
        self.color_true = '00FF00'
        self.color_false = 'FC371B'
        self.color_detail_wsdl = '94B4FF'
        self.color_detail_peoplesoft = 'D98FF5'
        self.wb = Workbook()


    def run(self, request):
        try:
            print("WorkerServiceComparisonExcel.run()")
            workers = self.worker_comparison.get_workers_comparison(request=request)

            workers_personal_data = workers['category']['personal_data']['data']
            workers_relationship_data = workers['category']['work_relationship']['data']
            workers_not_found = workers['not_found']['data']

            self.create_sheet_personal_data(workers_personal_data)
            self.create_sheet_relationship(workers_relationship_data)
            self.create_sheet_not_found(workers_not_found)

            # Crear un objeto de BytesIO para almacenar el archivo en memoria
            excel_file = BytesIO()

            # Guardar el libro de Excel en el objeto BytesIO
            self.wb.save(excel_file)

            # Configurar la respuesta HTTP para devolver el archivo Excel
            response = HttpResponse(
                excel_file.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )

            now = datetime.now()
            date_time = now.strftime("%d-%m-%Y_%H%M%S")
            print(date_time)
            file_name = f"workers_{date_time}.xlsx"

            # Configurar el encabezado de la respuesta HTTP
            response['Content-Disposition'] = 'attachment; filename=' + file_name
            
            return response
        except Exception as e:
            raise Exception(f" Location: def_run () | Error: {e}") from e

    def create_sheet_personal_data(self, personal_data):
        try:
            # Crear un libro de Excel y agregar una hoja de trabajo
            ws = self.wb.active

            # Cambiar el título de la hoja de trabajo
            ws.title = 'Workers'

            ws.append([
                'Person Number',
                'Nombre',
                'Nombre Detalle Hcm',
                'Nombre Detalle Peoplesoft',
                'Correo',
                'Correo Detalle Hcm',
                'Correo Detalle Peoplesoft',
                'Direccion',
                'Direccion Detalle Hcm',
                'Direccion Detalle Peoplesoft',
                'Direccion 2',
                'Direccion 2 Detalle Hcm',
                'Direccion 2 Detalle Peoplesoft',
                'Ciudad',
                'Ciudad Detalle Hcm',
                'Ciudad Detalle Peoplesoft'
            ])
            

            for idx, worker in enumerate(personal_data, start=2):
                ws.append([
                    worker['person_number'], #1
                    worker['name'], #2
                    worker['name_wsdl'], #3
                    worker['name_peoplesoft'], #4
                    worker['email'], #5
                    worker['email_wsdl'], #6
                    worker['email_peoplesoft'], #7
                    worker['address_1'], #8
                    worker['address_1_wsdl'], #9
                    worker['address_1_peoplesoft'], #10
                    worker['address_2'], #11
                    worker['address_2_wsdl'], #12
                    worker['address_2_peoplesoft'], #13
                    worker['city'], #14
                    worker['city_wsdl'], #15
                    worker['city_peoplesoft'] #16
                ])

                # Agregar color a las celdas

                # Nombre
                if worker['name'] == True:
                    cell_color = ws.cell(row=idx, column=2)
                    cell_color.fill = PatternFill(start_color=self.color_true, end_color=self.color_true, fill_type='solid')
                else:
                    cell_color = ws.cell(row=idx, column=2)
                    cell_color.fill = PatternFill(start_color=self.color_false, end_color=self.color_false, fill_type='solid')
                
                cell_color = ws.cell(row=idx, column=3)
                cell_color.fill = PatternFill(start_color=self.color_detail_wsdl, end_color=self.color_detail_wsdl, fill_type='solid')
                cell_color = ws.cell(row=idx, column=4)
                cell_color.fill = PatternFill(start_color=self.color_detail_peoplesoft, end_color=self.color_detail_peoplesoft, fill_type='solid')


                # Email
                if worker['email'] == True: 
                    cell_color = ws.cell(row=idx, column=5)
                    cell_color.fill = PatternFill(start_color=self.color_true, end_color=self.color_true, fill_type='solid')
                else:
                    cell_color = ws.cell(row=idx, column=5)
                    cell_color.fill = PatternFill(start_color=self.color_false, end_color=self.color_false, fill_type='solid')

                cell_color = ws.cell(row=idx, column=6)
                cell_color.fill = PatternFill(start_color=self.color_detail_wsdl, end_color=self.color_detail_wsdl, fill_type='solid')
                cell_color = ws.cell(row=idx, column=7)
                cell_color.fill = PatternFill(start_color=self.color_detail_peoplesoft, end_color=self.color_detail_peoplesoft, fill_type='solid')


                # Address                
                if worker['address_1'] == True:
                    cell_color = ws.cell(row=idx, column=8)
                    cell_color.fill = PatternFill(start_color=self.color_true, end_color=self.color_true, fill_type='solid')
                else:
                    cell_color = ws.cell(row=idx, column=8)
                    cell_color.fill = PatternFill(start_color=self.color_false, end_color=self.color_false, fill_type='solid')

                cell_color = ws.cell(row=idx, column=9)
                cell_color.fill = PatternFill(start_color=self.color_detail_wsdl, end_color=self.color_detail_wsdl, fill_type='solid')
                cell_color = ws.cell(row=idx, column=10)
                cell_color.fill = PatternFill(start_color=self.color_detail_peoplesoft, end_color=self.color_detail_peoplesoft, fill_type='solid')

                # Address 2
                if worker['address_2'] == True:
                    cell_color = ws.cell(row=idx, column=11)
                    cell_color.fill = PatternFill(start_color=self.color_true, end_color=self.color_true, fill_type='solid')
                else:
                    cell_color = ws.cell(row=idx, column=11)
                    cell_color.fill = PatternFill(start_color=self.color_false, end_color=self.color_false, fill_type='solid')

                cell_color = ws.cell(row=idx, column=12)
                cell_color.fill = PatternFill(start_color=self.color_detail_wsdl, end_color=self.color_detail_wsdl, fill_type='solid')
                cell_color = ws.cell(row=idx, column=13)
                cell_color.fill = PatternFill(start_color=self.color_detail_peoplesoft, end_color=self.color_detail_peoplesoft, fill_type='solid')

                # City
                if worker['city'] == True:
                    cell_color = ws.cell(row=idx, column=14)
                    cell_color.fill = PatternFill(start_color=self.color_true, end_color=self.color_true, fill_type='solid')
                else:
                    cell_color = ws.cell(row=idx, column=14)
                    cell_color.fill = PatternFill(start_color=self.color_false, end_color=self.color_false, fill_type='solid')
                
                cell_color = ws.cell(row=idx, column=15)
                cell_color.fill = PatternFill(start_color=self.color_detail_wsdl, end_color=self.color_detail_wsdl, fill_type='solid')
                cell_color = ws.cell(row=idx, column=16)
                cell_color.fill = PatternFill(start_color=self.color_detail_peoplesoft, end_color=self.color_detail_peoplesoft, fill_type='solid')
            
        except Exception as e:
            raise Exception(f" Location: def_crear_hoja_personal_data () | Error: {e}") from e

    def create_sheet_not_found(self, not_found):
        try:
            ws = self.wb.create_sheet(title='Not Found')
            ws.append([
                'Person Number'
            ])

            for idx, worker in enumerate(not_found, start=2):
                ws.append([
                    worker['person_number']
                ])
        
        except Exception as e:
            raise Exception(f" Location: def_crear_hoja_not_found () | Error: {e}") from e

    def create_sheet_relationship(self, relationship_data):
        try:
            ws = self.wb.create_sheet(title='Relationships')
            ws.append([
                'Person Number',
                'Location Code',
                'Location Code Detalle Hcm',
                'Location Code Detalle Peoplesoft',
                'Codigo centro de costo',
                'Codigo centro de costo Detalle Hcm',
                'Codigo centro de costo Detalle Peoplesoft',
                'Job Name',
                'Job Name Detalle Hcm',
                'Job Name Detalle Peoplesoft',
                'Job Code',
                'Job Code Detalle Hcm',
                'Job Code Detalle Peoplesoft',
                'Manager Id',
                'Manager Id Detalle Hcm',
                'Manager Id Detalle Peoplesoft'
            ])

            for idx, worker in enumerate(relationship_data, start=2):
                ws.append([
                    worker['person_number'],
                    worker['location_code'],
                    worker['location_code_wsdl'],
                    worker['location_code_peoplesoft'],
                    worker['codigo_centro_costo'],
                    worker['codigo_centro_costo_wsdl'],
                    worker['codigo_centro_costo_peoplesoft'],
                    worker['job_name'],
                    worker['job_name_wsdl'],
                    worker['job_name_peoplesoft'],
                    worker['job_code'],
                    worker['job_code_wsdl'],
                    worker['job_code_peoplesoft'],
                    worker['manager_id'],
                    worker['manager_id_wsdl'],
                    worker['manager_id_peoplesoft']
                ])

                # Agregar color a las celdas

                # Location Code
                if worker['location_code'] == True:
                    cell_color = ws.cell(row=idx, column=2)
                    cell_color.fill = PatternFill(start_color=self.color_true, end_color=self.color_true, fill_type='solid')
                else:
                    cell_color = ws.cell(row=idx, column=2)
                    cell_color.fill = PatternFill(start_color=self.color_false, end_color=self.color_false, fill_type='solid')

                cell_color = ws.cell(row=idx, column=3)
                cell_color.fill = PatternFill(start_color=self.color_detail_wsdl, end_color=self.color_detail_wsdl, fill_type='solid')
                cell_color = ws.cell(row=idx, column=4)
                cell_color.fill = PatternFill(start_color=self.color_detail_peoplesoft, end_color=self.color_detail_peoplesoft, fill_type='solid')

                # Codigo centro de costo
                if worker['codigo_centro_costo'] == True:
                    cell_color = ws.cell(row=idx, column=5)
                    cell_color.fill = PatternFill(start_color=self.color_true, end_color=self.color_true, fill_type='solid')
                else:
                    cell_color = ws.cell(row=idx, column=5)
                    cell_color.fill = PatternFill(start_color=self.color_false, end_color=self.color_false, fill_type='solid')

                cell_color = ws.cell(row=idx, column=6)
                cell_color.fill = PatternFill(start_color=self.color_detail_wsdl, end_color=self.color_detail_wsdl, fill_type='solid')
                cell_color = ws.cell(row=idx, column=7)
                cell_color.fill = PatternFill(start_color=self.color_detail_peoplesoft, end_color=self.color_detail_peoplesoft, fill_type='solid')

                # Job Name
                if worker['job_name'] == True:
                    cell_color = ws.cell(row=idx, column=8)
                    cell_color.fill = PatternFill(start_color=self.color_true, end_color=self.color_true, fill_type='solid')
                else:
                    cell_color = ws.cell(row=idx, column=8)
                    cell_color.fill = PatternFill(start_color=self.color_false, end_color=self.color_false, fill_type='solid')
                
                cell_color = ws.cell(row=idx, column=9)
                cell_color.fill = PatternFill(start_color=self.color_detail_wsdl, end_color=self.color_detail_wsdl, fill_type='solid')
                cell_color = ws.cell(row=idx, column=10)
                cell_color.fill = PatternFill(start_color=self.color_detail_peoplesoft, end_color=self.color_detail_peoplesoft, fill_type='solid')

                # Job Code
                if worker['job_code'] == True:
                    cell_color = ws.cell(row=idx, column=11)
                    cell_color.fill = PatternFill(start_color=self.color_true, end_color=self.color_true, fill_type='solid')
                else:
                    cell_color = ws.cell(row=idx, column=11)
                    cell_color.fill = PatternFill(start_color=self.color_false, end_color=self.color_false, fill_type='solid')
                
                cell_color = ws.cell(row=idx, column=12)
                cell_color.fill = PatternFill(start_color=self.color_detail_wsdl, end_color=self.color_detail_wsdl, fill_type='solid')
                cell_color = ws.cell(row=idx, column=13)
                cell_color.fill = PatternFill(start_color=self.color_detail_peoplesoft, end_color=self.color_detail_peoplesoft, fill_type='solid')

                # Manager Id
                if worker['manager_id'] == True:
                    cell_color = ws.cell(row=idx, column=14)
                    cell_color.fill = PatternFill(start_color=self.color_true, end_color=self.color_true, fill_type='solid')
                else:
                    cell_color = ws.cell(row=idx, column=14)
                    cell_color.fill = PatternFill(start_color=self.color_false, end_color=self.color_false, fill_type='solid')

                cell_color = ws.cell(row=idx, column=15)
                cell_color.fill = PatternFill(start_color=self.color_detail_wsdl, end_color=self.color_detail_wsdl, fill_type='solid')
                cell_color = ws.cell(row=idx, column=16)
                cell_color.fill = PatternFill(start_color=self.color_detail_peoplesoft, end_color=self.color_detail_peoplesoft, fill_type='solid')
            
        except Exception as e:
            raise Exception(f" Location: def_crear_hoja_relationship () | Error: {e}") from e
        

    def create_sheet_spreadsheet_email(self, personal_data):
        try:
            ws = self.wb.create_sheet(title='Spreadsheet Email')
        except Exception as e:
            raise Exception(f" Location: def_create_sheet_spreadsheet_email () | Error: {e}") from e

